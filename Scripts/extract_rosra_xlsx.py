"""Extract ROSRA v4 source-of-truth workbook into a structured JSON dump.

The workbook is the authoritative reference for ROSRA business logic. We dump
every sheet's cells (values + formulas) so the assistant can use the file as
the canonical source for revenue-gap formulas, defaults, and structure.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def cell_repr(cell):
    val = cell.value
    if val is None:
        return None
    # If a formula, return both formula and any cached value if present.
    if isinstance(val, str) and val.startswith("="):
        return {"f": val}
    return val


def extract(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=False)
    out = {
        "file": str(xlsx_path),
        "sheets": [],
    }
    for ws in wb.worksheets:
        sheet = {
            "name": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "dimensions": ws.dimensions,
            "merged_cells": [str(r) for r in ws.merged_cells.ranges],
            "cells": {},
        }
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell_repr(cell)
                if v is None:
                    continue
                addr = f"{get_column_letter(cell.column)}{cell.row}"
                sheet["cells"][addr] = v
        out["sheets"].append(sheet)
    return out


def main():
    src = Path(sys.argv[1])
    dst = Path(sys.argv[2])
    data = extract(src)
    dst.write_text(json.dumps(data, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"Sheets: {[s['name'] for s in data['sheets']]}")
    for s in data["sheets"]:
        print(f"  {s['name']}: {s['max_row']} rows x {s['max_col']} cols, {len(s['cells'])} non-empty cells")


if __name__ == "__main__":
    main()
