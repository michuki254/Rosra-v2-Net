"""Generate the v13 top-down analysis seed files from the two workbooks
Omar circulated on 2026-06-10:

  Data/Wofi/wofi_gni_atlas.json      124 rows  (iso3, country, gni_per_capita_atlas_usd,
                                                data_year, source) from GNI_Filter_Data.
                                                Used ONLY for staged peer filtering.
  Data/Wofi/wofi_assumptions.json    1 object  staged-peer-filter + frontier constants
                                                from the Assumptions sheet.
  Data/SeedData/peersng.json         184 rows  preloaded domestic peer SNG data
                                                (11 countries, USD, 2024) from
                                                "osr and subnational gdp for some countries.xlsx".

Usage (repo root):
  python tools/scripts/generate_wofi_v13_seed_files.py \
      "C:/Users/dxmic/Downloads/ROSRA updated quick osr assessment_global.xlsx" \
      "C:/Users/dxmic/Downloads/osr and subnational gdp for some countries.xlsx"

Requires openpyxl.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]

COUNTRY_ISO3 = {
    "Kenya": "KEN",
    "India": "IND",
    "Nigeria": "NGA",
    "Philippines": "PHL",
    "Morocco": "MAR",
    "South Africa": "ZAF",
    "Sri Lanka": "LKA",
    "Mexico": "MEX",
    "Colombia": "COL",
    "Indonesia": "IDN",
}


def build_gni_atlas(wb) -> list[dict]:
    ws = wb["GNI_Filter_Data"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        iso3, country, _income, _proxy, _mult, gni, gni_year, source = row[:8]
        if iso3 is None:
            continue
        out.append(
            {
                "iso3": iso3,
                "country": country,
                "gni_per_capita_atlas_usd": gni,
                "data_year": gni_year,
                "source": source,
            }
        )
    return out


def build_assumptions(wb) -> dict:
    ws = wb["Assumptions"]
    vals = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    return {
        "min_peer_count": vals["Minimum peer count before widening"],
        "stage_a_gni_lower": vals["Stage A GNI ratio lower"],
        "stage_a_gni_upper": vals["Stage A GNI ratio upper"],
        "stage_b_gni_lower": vals["Stage B GNI ratio lower"],
        "stage_b_gni_upper": vals["Stage B GNI ratio upper"],
        "stage_d_nearest_n": vals["Nearest-neighbour fallback count"],
        "base_frontier_top_share": 0.20,
        "internal_dispersion_factor": vals["Internal-dispersion adjustment"],
        "total_osr_uplift": vals["Total-OSR uplift"],
        "source": "ROSRA updated quick osr assessment_global.xlsx (v13), Assumptions sheet",
    }


def build_peersng(wb) -> list[dict]:
    ws = wb["Sheet1"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        peer_id, country, sng, band, gcp, osr = row[:6]
        watchlist = row[7]
        if peer_id is None:
            continue
        iso3 = COUNTRY_ISO3[country]
        out.append(
            {
                "CountryCode": iso3,
                "Country": country,
                "SNG": sng,
                "Band": band,
                "OSR": round(float(osr), 2),
                "GCP": round(float(gcp), 2),
                "Population": 0,
                "Include": True,
                "Watchlist": bool(watchlist),
                "Currency": "USD",
                "DataYear": 2024,
            }
        )
    return out


def main() -> None:
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    global_wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
    peers_wb = openpyxl.load_workbook(sys.argv[2], data_only=True)

    targets = {
        REPO_ROOT / "Data/Wofi/wofi_gni_atlas.json": build_gni_atlas(global_wb),
        REPO_ROOT / "Data/Wofi/wofi_assumptions.json": build_assumptions(global_wb),
        REPO_ROOT / "Data/SeedData/peersng.json": build_peersng(peers_wb),
    }
    for path, payload in targets.items():
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        count = len(payload) if isinstance(payload, list) else 1
        print(f"wrote {path.relative_to(REPO_ROOT)}  ({count} records)")


if __name__ == "__main__":
    main()
